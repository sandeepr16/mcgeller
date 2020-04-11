from flask import Flask, render_template, request
app = Flask(__name__)
import random
import openpyxl
from flask import *  
app = Flask(__name__)  

  
@app.route('/utg',methods = ['POST'])

def utg():  
    
    loc = ('C:/Users/saira/Desktop/utg/utg.xlsx')
    wbr = openpyxl.load_workbook(loc)
    s1 = wbr.worksheets[0]
    wbr.create_sheet(index = 1 , title = 'ctt')
    wbr.create_sheet(index = 2 , title = 'ttt')
    s2 = wbr.worksheets[1]
    s3 = wbr.worksheets[2]
    cl = ''
    val = ''
    te = ''
    telis = []
    cllis = []
    for i in range(1,s1.max_row+1):
        telis.append(s1.cell(row = i,column = 3).value)
    telis = list(dict.fromkeys(telis))
    for i in range(1,s1.max_row+1):
        cllis.append(s1.cell(row = i,column = 1).value)
    cllis = list(dict.fromkeys(cllis))
    g = 1
    for a in range(len(cllis)):
        cl = cllis[a]
        for i in range(7):
        
            if(g%7 == 1):
                for j in range(1,10):
                    if(j==1):
                        val = cl
                    elif(j==2):
                        val = '1'
                    elif(j==3):
                        val = '2'
                    elif(j==4):
                        val = '3'
                    elif(j==5):
                        val = '4'
                    elif(j==6):
                        val = 'lunch'
                    elif(j==7):
                        val = '5'
                    elif(j==8):
                        val = '6'
                    elif(j==9):
                        val = '7'
                    else:
                        break
                    s2.cell(row = g,column = j).value = val
                g = g+1
            else:
                k = g%7
                if(k==2):
                    val = 'mon'
                if(k==3):
                    val = 'tue'
                if(k==4):
                    val = 'wed'
                if(k==5):
                    val = 'thu'
                if(k==6):
                    val = 'fri'
                if(k==0):
                    val = 'sat'
                    s2.cell(row = g,column = 1).value = val
                    g = g+1
                    break
                s2.cell(row = g,column = 1).value = val
                g = g+1
    g = 1
    for a in range(len(telis)):
        te = telis[a]
        for i in range(7):
            if(g%7 == 1):
                for j in range(1,10):
                    if(j==1):
                        val = te
                    elif(j==2):
                        val = '1'
                    elif(j==3):
                        val = '2'
                    elif(j==4):
                        val = '3'
                    elif(j==5):
                        val = '4'
                    elif(j==6):
                        val = 'lunch'
                    elif(j==7):
                        val = '5'
                    elif(j==8):
                        val = '6'
                    elif(j==9):
                        val = '7'
                    else:
                        break
                    s3.cell(row = g,column = j).value = val
                g = g+1
            else:
                k = g%7
                if(k==2):
                    val = 'mon'
                if(k==3):
                    val = 'tue'
                if(k==4):
                    val = 'wed'
                if(k==5):
                    val = 'thu'
                if(k==6):
                    val = 'fri'
                if(k==0):
                    val = 'sat'
                    s3.cell(row = g,column = 1).value = val
                    g = g+1
                    break
                s3.cell(row = g,column = 1).value = val
                g = g+1
    ch1 = []
    ch2 = []
    choices = list(range(1,s1.max_row+1))
    for i in range(len(choices)):
        r = choices[i]
        subject = s1.cell(row = r,column = 2).value
        if(subject.find('LAB') != -1):
            ch1.append(choices[i])
        else:
            ch2.append(choices[i])
    random.shuffle(ch1)
    random.shuffle(ch2)
    w = 0
    x = 0
    for b in range(len(ch1)):
        r = ch1[b]
        clsname = s1.cell(row = r,column = 1).value
        subject = s1.cell(row = r,column = 2).value
        teacher = s1.cell(row = r,column = 3).value
        noofcls = s1.cell(row = r,column = 4).value
        for y in range(1,s2.max_row+1):
            if(s2.cell(row = y,column = 1).value == clsname):
                w = y
            y=y+7;    
        for y in range(1,s3.max_row+1):
            if(s3.cell(row = y,column = 1).value == teacher):
                x = y
            y=y+7;    
        while True:
            i = random.randrange(1,7,1)
            j = random.choice([2,3,7])
            if (s2.cell(row = w+i,column = j).value == None and s2.cell(row = w+i,column = j+1).value == None and s2.cell(row = w+i,column = j+2).value == None and s3.cell(row = x+i,column = j).value == None and s3.cell(row = x+i,column = j+1).value == None and s3.cell(row = x+i,column = j+2).value == None):
                s2.cell(row = w+i,column = j).value = subject
                s2.cell(row = w+i,column = j+1).value = subject
                s2.cell(row = w+i,column = j+2).value = subject
                s3.cell(row = x+i,column = j).value = clsname
                s3.cell(row = x+i,column = j+1).value = clsname
                s3.cell(row = x+i,column = j+2).value = clsname
                break
            else:
                continue
    br = 0
    for b in range(len(ch2)):
        r = ch2[b]
        clsname = s1.cell(row = r,column = 1).value
        subject = s1.cell(row = r,column = 2).value
        teacher = s1.cell(row = r,column = 3).value
        noofcls = s1.cell(row = r,column = 4).value
        for y in range(1,s2.max_row+1):
            if(s2.cell(row = y,column = 1).value == clsname):
                w = y
            y=y+7;    
        for y in range(1,s3.max_row+1):
            if(s3.cell(row = y,column = 1).value == teacher):
                x = y
            y=y+7;    
        for k in range(int(noofcls)):
            while True:
                br = br+1
                i = random.randrange(1,7,1)
                j = random.randrange(2,10,1)
                if ( j == 6 ):
                    continue
                elif(s2.cell(row = w+i,column = j).value == None and s3.cell(row = x+i,column = j).value == None):
                    s2.cell(row = w+i,column = j).value = subject
                    s3.cell(row = x+i,column = j).value = clsname
                    break
                elif(br>10000):
                    break
                else:
                    continue
    wbr.save('C:/Users/saira/Desktop/utg/utg1.xlsx')
    return render_template('kk.html')
if __name__ == '__main__':
     app.run()  